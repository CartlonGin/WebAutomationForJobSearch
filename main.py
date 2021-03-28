from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import time
from selenium.common import exceptions
import openpyxl




def delete(sheet):
    # continuously delete row 2 until there
    # is only a single row left over
    # that contains column names
    while (sheet.max_row > 1):
        # this method removes the row 2
        sheet.delete_rows(2)
    # return to main function
    return

pathExcel = r"C:\Users\kaanh\Desktop\job.xlsx"
workbook = openpyxl.load_workbook(pathExcel)
sheet = workbook.active
delete(sheet)
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)

driver.get("https://www.workopolis.com/en/")

searchTitle = driver.find_element_by_id("keyword")
searchTitle.send_keys("software developer")
searchTitle.send_keys(Keys.TAB)


searchArea = driver.find_element_by_id("location")
searchArea.send_keys("toronto")
searchArea.send_keys(Keys.RETURN)

driver.implicitly_wait(5)


r = 1
for pageNumber in range(1, 30):
    nextPage = driver.find_element_by_link_text(str(pageNumber)).click()
    for i in range(20):
        jobCard = driver.find_elements_by_class_name("JobCard")
        jobCard[i].click()
        driver.implicitly_wait(10)
        exist = True
        try:
            jobSkills = driver.find_elements_by_class_name("nav")
            exist = True

        except NoSuchElementException:
            exist = False
        if exist:
            for j in range(len(jobSkills)):
                try:
                    if jobSkills[j].text.__contains__("Python") or jobSkills[j].text.__contains__("Java"):
                        print("found!")
                        driver.implicitly_wait(10)
                        jobTitle = driver.find_element_by_class_name("ViewJobHeader-title")
                        print(jobTitle.text)
                        company = driver.find_element_by_class_name("ViewJobHeader-company")
                        print(company.text)
                        salary = driver.find_element_by_class_name("Salary")
                        print(salary.text)
                        jobDescripton = driver.find_element_by_css_selector("html[data-css~=serp] .ViewJob-description")
                        sheet.cell(row=r, column=1).value = jobTitle.text
                        sheet.cell(row=r, column=2).value = company.text
                        sheet.cell(row=r, column=3).value = salary.text
                        sheet.cell(row=r, column=4).value = jobDescripton.text
                        r = r+1

                except exceptions.StaleElementReferenceException:

                    pass

workbook.save(pathExcel)

time.sleep(5)
driver.quit()
